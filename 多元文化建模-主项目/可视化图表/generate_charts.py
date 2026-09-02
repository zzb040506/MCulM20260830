# -*- coding: utf-8 -*-
"""
多元文化建模项目可视化图表生成脚本
生成9张图：4张统计雷达图 + 2张补充统计图 + 3张设计叙事流程图
全部使用项目Excel真实数据，中英双语标签
"""
import os
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
from matplotlib.lines import Line2D

# ---------------- 全局配置 ----------------
BASE = "/Users/f.fantasiachopin/Documents/UCAS博士文件夹/Project/多元文化建模20260830/多元文化建模-主项目"
OUT = os.path.join(BASE, "可视化图表")
os.makedirs(OUT, exist_ok=True)

plt.rcParams["font.sans-serif"] = ["PingFang SC", "Arial Unicode MS", "Heiti TC"]
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["figure.dpi"] = 150
plt.rcParams["savefig.dpi"] = 200

# 配色（学术风，低饱和）
C_MAIN = "#2E5C8A"   # 主蓝
C_ACC  = "#C0504D"   # 强调红
C_GREY = "#7F7F7F"
C_GREEN = "#4F8A5B"
C_ORANGE = "#D98A2B"
PALETTE = ["#2E5C8A", "#C0504D", "#4F8A5B", "#D98A2B", "#7F6BAF", "#6A8FAF"]

# ---------------- 数据读取 ----------------
def load_wvs():
    wb = openpyxl.load_workbook(os.path.join(BASE, "国家文化知识建模_wvs编码结果表.xlsx"), data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows

def load_theory():
    wb = openpyxl.load_workbook(os.path.join(BASE, "国家文化知识建模_中层理论注册表.xlsx"), data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows

def load_semantic():
    wb = openpyxl.load_workbook(os.path.join(BASE, "国家文化知识建模_语义编号表.xlsx"), data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows

WVS = load_wvs()
THEORY = load_theory()
SEM = load_semantic()
N_Q = len(WVS)
N_T = len(THEORY)

# 代码中文映射（来自语义编号表）
J_ZH = {"J-CM":"分类/成员","J-AC":"属性/因果","J-ER":"评价排序","J-NO":"规范/应然",
        "J-RA":"关系/分配","J-RI":"表征/认同","—":"未编码(元数据)"}
D_ZH = {"D-FI":"家庭","D-CR":"社区","D-RS":"宗教","D-PL":"政治法律","D-WO":"市场工作",
        "D-KT":"教育知识","D-MC":"媒介文化","D-HC":"健康身体","—":"未编码(元数据)"}
O_ZH = {"O-GC":"性别","O-SF":"性取向","O-AG":"年龄世代","O-SP":"阶层","O-RE":"种族",
        "O-NM":"国族移民","O-RS":"宗教归属","O-LG":"语言","O-BH":"身体健康","O-FR":"亲属角色",
        "O-ST":"关系距离","O-AR":"专业角色","O-GP":"群体运动","O-OI":"组织机构","O-PL":"行为实践",
        "O-RP":"规范制度","O-KT":"知识技术","O-MSW":"文本叙事","O-OSN":"物景生态","—":"未编码"}

# ---------------- 雷达图通用函数 ----------------
def radar(ax, labels, values, color, label, alpha=0.25):
    n = len(labels)
    angles = np.linspace(0, 2*np.pi, n, endpoint=False).tolist()
    vc = list(values) + [values[0]]
    ac = angles + [angles[0]]
    ax.plot(ac, vc, color=color, lw=2, label=label)
    ax.fill(ac, vc, color=color, alpha=alpha)
    ax.set_xticks(angles)
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    ax.grid(True, color="#CCCCCC", lw=0.6)

# ================= 图1：290题×判断类型J =================
def fig01():
    from collections import Counter
    cnt = Counter(r["J_code"] for r in WVS)
    order = ["J-CM","J-AC","J-ER","J-NO","J-RA","J-RI"]
    labels = [f"{J_ZH[c]}\n{c}" for c in order]
    vals = [cnt.get(c,0) for c in order]
    n_meta = cnt.get("—",0)

    fig, ax = plt.subplots(figsize=(8.5,8), subplot_kw=dict(projection="polar"))
    radar(ax, labels, vals, C_MAIN, f"内容判断题项 Content-coded (n={sum(vals)})")
    ax.set_title(f"图1  290题 × 判断类型 J 分布\nFig.1  290 Items × Judgment Type J",
                 fontsize=13, pad=28, fontweight="bold")
    ax.set_rlabel_position(90)
    ax.text(0.5, -0.06, f"注：另有 {n_meta} 题为元数据/结果型变量(—)，不进入内容判断编码\n"
            f"Note: {n_meta} items are metadata/outcome (—), excluded from content coding",
            transform=ax.transAxes, ha="center", fontsize=8.5, color=C_GREY)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig01_judgment_J.png"), bbox_inches="tight")
    plt.close()

# ================= 图2：290题×场域D =================
def fig02():
    from collections import Counter
    cnt = Counter(r["D_primary"] for r in WVS)
    order = ["D-FI","D-CR","D-RS","D-PL","D-WO","D-KT","D-MC","D-HC"]
    labels = [f"{D_ZH[c]}\n{c}" for c in order]
    vals = [cnt.get(c,0) for c in order]
    n_meta = cnt.get("—",0)

    fig, ax = plt.subplots(figsize=(8.5,8), subplot_kw=dict(projection="polar"))
    radar(ax, labels, vals, C_GREEN, f"内容场域题项 Content-domain (n={sum(vals)})")
    ax.set_title(f"图2  290题 × 制度与生活场域 D 分布\nFig.2  290 Items × Institutional Domain D",
                 fontsize=13, pad=28, fontweight="bold")
    ax.set_rlabel_position(90)
    ax.text(0.5, -0.06, f"注：另有 {n_meta} 题为元数据/结果型变量(—)，不进入场域编码\n"
            f"Note: {n_meta} items are metadata/outcome (—)",
            transform=ax.transAxes, ha="center", fontsize=8.5, color=C_GREY)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig02_domain_D.png"), bbox_inches="tight")
    plt.close()

# ================= 图3：290题×对象O（重点类） =================
def fig03():
    from collections import Counter
    cnt = Counter(r["O_primary"] for r in WVS)
    # 取出现>=3次或重点类
    focus = ["O-PL","O-RP","O-OI","O-ST","O-NM","O-RS","O-AR","O-GC","O-FR","O-KT","O-GP","O-MSW","O-SP"]
    focus = [c for c in focus if cnt.get(c,0) > 0]
    order = sorted(focus, key=lambda c: -cnt.get(c,0))
    labels = [f"{O_ZH.get(c,c)}\n{c}" for c in order]
    vals = [cnt.get(c,0) for c in order]
    n_meta = cnt.get("—",0)

    fig, ax = plt.subplots(figsize=(9.5,8.5), subplot_kw=dict(projection="polar"))
    radar(ax, labels, vals, C_ORANGE, f"判断对象 Object O (n={sum(vals)})")
    ax.set_title(f"图3  290题 × 文化判断对象 O 分布\nFig.3  290 Items × Cultural Object O",
                 fontsize=13, pad=28, fontweight="bold")
    ax.set_rlabel_position(90)
    ax.text(0.5, -0.06, f"注：另有 {n_meta} 题为元数据/结果型变量(—)\nNote: {n_meta} items are metadata/outcome (—)",
            transform=ax.transAxes, ha="center", fontsize=8.5, color=C_GREY)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig03_object_O.png"), bbox_inches="tight")
    plt.close()

# ================= 图4：34理论×板块section =================
def fig04():
    from collections import Counter
    cnt = Counter(r["section"] for r in THEORY)
    # 简短名
    short = {
        "跨场域的文化组织机制":"跨场域机制\nCross-domain",
        "家庭、关系与地方共同体":"家庭与共同体\nFamily & Community",
        "分层、性别与交叉身份":"分层与身份\nStratification & Identity",
        "宗教、政治、法律、国族与迁移":"宗教政治国族\nReligion/Polity/Nation",
        "市场、组织、职业、知识与科学":"市场知识与科学\nMarket/Knowledge/Science",
        "健康、身体、照料、记忆与创伤":"健康身体记忆\nHealth/Body/Memory",
    }
    order = ["跨场域的文化组织机制","家庭、关系与地方共同体","分层、性别与交叉身份",
             "宗教、政治、法律、国族与迁移","市场、组织、职业、知识与科学","健康、身体、照料、记忆与创伤"]
    labels = [short[c] for c in order]
    vals = [cnt.get(c,0) for c in order]

    fig, ax = plt.subplots(figsize=(9,8.5), subplot_kw=dict(projection="polar"))
    radar(ax, labels, vals, C_ACC, f"中层理论条目 Mid-level theories (n={N_T})")
    ax.set_title(f"图4  34条中层理论 × 板块分布\nFig.4  34 Mid-level Theories × Section",
                 fontsize=13, pad=28, fontweight="bold")
    ax.set_rlabel_position(90)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig04_theory_section.png"), bbox_inches="tight")
    plt.close()

# ================= 图5：review_status 分布 =================
def fig05():
    from collections import Counter
    cnt = Counter(r["review_status"] for r in WVS)
    items = sorted(cnt.items(), key=lambda x: -x[1])
    labels = [k for k,_ in items]
    vals = [v for _,v in items]

    # 颜色：PASS类绿，ADJUDICATED类红，其他灰
    colors = []
    for k in labels:
        if k == "PASS": colors.append(C_GREEN)
        elif k.startswith("ADJUDICATED"): colors.append(C_ACC)
        elif k.startswith("PASS"): colors.append(C_ORANGE)
        else: colors.append(C_GREY)

    fig, ax = plt.subplots(figsize=(11,7))
    y = np.arange(len(labels))
    bars = ax.barh(y, vals, color=colors, edgecolor="white")
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("题项数 Number of items", fontsize=10)
    ax.set_title(f"图5  290题 × 复核裁决状态 review_status 分布\n"
                 f"Fig.5  290 Items × Review/Adjudication Status", fontsize=12.5, fontweight="bold", pad=14)
    for bar, v in zip(bars, vals):
        ax.text(bar.get_width()+0.5, bar.get_y()+bar.get_height()/2, str(v), va="center", fontsize=8.5)
    ax.text(0.99, -0.12, "绿色=直接通过 PASS  红色=经裁决 ADJUDICATED  橙色=带限制通过 PASS_*  → 体现编码的边界意识与审计痕迹\n"
            "Green=PASS  Red=Adjudicated  Orange=Pass-with-limit  → shows boundary-aware, audited coding",
            transform=ax.transAxes, ha="right", fontsize=8, color=C_GREY)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig05_review_status.png"), bbox_inches="tight")
    plt.close()

# ================= 图6：J × D 交叉热力图 =================
def fig06():
    from collections import Counter
    j_order = ["J-CM","J-AC","J-ER","J-NO","J-RA","J-RI"]
    d_order = ["D-FI","D-CR","D-RS","D-PL","D-WO","D-KT","D-MC","D-HC"]
    mat = np.zeros((len(j_order), len(d_order)), dtype=int)
    for r in WVS:
        j = r["J_code"]; d = r["D_primary"]
        if j in j_order and d in d_order:
            mat[j_order.index(j), d_order.index(d)] += 1

    fig, ax = plt.subplots(figsize=(11,7.5))
    im = ax.imshow(mat, cmap="Blues", aspect="auto")
    ax.set_xticks(range(len(d_order)))
    ax.set_xticklabels([f"{D_ZH[d]}\n{d}" for d in d_order], fontsize=9)
    ax.set_yticks(range(len(j_order)))
    ax.set_yticklabels([f"{J_ZH[j]} ({j})" for j in j_order], fontsize=9)
    ax.set_xlabel("制度与生活场域 Domain D", fontsize=10)
    ax.set_ylabel("判断类型 Judgment J", fontsize=10)
    ax.set_title(f"图6  判断类型 J × 场域 D 交叉分布热力图\n"
                 f"Fig.6  Judgment × Domain Crosstab Heatmap (content-coded items)",
                 fontsize=12.5, fontweight="bold", pad=14)
    # 标注数字
    vmax = mat.max()
    for i in range(len(j_order)):
        for j in range(len(d_order)):
            v = mat[i,j]
            if v > 0:
                color = "white" if v > vmax*0.5 else "#333333"
                ax.text(j, i, str(v), ha="center", va="center", fontsize=9, color=color, fontweight="bold")
    fig.colorbar(im, ax=ax, label="题项数 Count", shrink=0.8)
    ax.text(0.5, -0.16, "注：不含101条元数据/结果型(—)题项。深色格=该判断在该场域高频出现，反映文化命题的结构性聚集\n"
            "Note: excludes 101 metadata (—) items. Darker cells show structural clustering of cultural propositions",
            transform=ax.transAxes, ha="center", fontsize=8, color=C_GREY)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig06_JxD_heatmap.png"), bbox_inches="tight")
    plt.close()

# ---------------- 流程图辅助 ----------------
def box(ax, x, y, w, h, text, fc, ec=None, tc="black", fs=10, weight="normal"):
    b = FancyBboxPatch((x, y), w, h, boxstyle="round,pad=0.02,rounding_size=0.08",
                       linewidth=1.2, facecolor=fc, edgecolor=ec or fc)
    ax.add_patch(b)
    ax.text(x+w/2, y+h/2, text, ha="center", va="center", fontsize=fs, color=tc, weight=weight, wrap=True)

def arrow(ax, x1, y1, x2, y2, color=C_GREY, lw=1.6, style="-|>"):
    a = FancyArrowPatch((x1,y1),(x2,y2), arrowstyle=style, mutation_scale=14,
                        color=color, lw=lw, connectionstyle="arc3,rad=0")
    ax.add_patch(a)

# ================= 图7：6层框架架构图 =================
def fig07():
    fig, ax = plt.subplots(figsize=(13,9))
    ax.set_xlim(0,13); ax.set_ylim(0,9.5); ax.axis("off")
    ax.set_title("图7  跨文化知识建模核心框架：6层维度如何组成命题与结构\n"
                "Fig.7  Core Framework: 6 Layers composing Propositions & Structures",
                fontsize=13, fontweight="bold", pad=10)

    # 左侧6层维度卡片
    layers = [
        ("C", "核心价值取向\nCore value orientation", "自主—嵌入 / 平等—等级 / 掌控—和谐", C_MAIN),
        ("D", "制度与生活场域\nDomain", "家庭/社区/宗教/政治/市场/教育/媒介/健康 8类", C_GREEN),
        ("A→O", "文化判断对象\nObject", "性别/阶层/国族/亲属/角色/规范/知识/叙事…19类", C_ORANGE),
        ("J", "判断与关系规则\nJudgment", "分类/属性因果/评价/规范/关系分配/表征 6类", C_ACC),
        ("S", "跨命题文化结构\nStructure", "边界/图式/框架/二元代码/叙事/话语场/体裁 7类", "#7F6BAF"),
        ("X→C", "实质情境条件\nContext", "资源/关系历史/规则/替代/严重性/问责/威胁/跨国…13类", C_GREY),
    ]
    y0 = 8.4
    for i,(code,name,desc,col) in enumerate(layers):
        y = y0 - i*1.25
        # 代码标签圆
        from matplotlib.patches import Circle
        c = Circle((0.55, y+0.32), 0.32, facecolor=col, edgecolor="white", lw=1.5, zorder=3)
        ax.add_patch(c)
        ax.text(0.55, y+0.32, code, ha="center", va="center", color="white", fontsize=9.5, fontweight="bold", zorder=4)
        box(ax, 1.1, y, 5.7, 0.62, f"{name}\n{desc}", "#FFFFFF", ec=col, fs=8.8)

    # 右侧：单题命题公式
    box(ax, 7.2, 6.6, 5.5, 1.1, "单一文化命题\nSingle Proposition\n$U_i = (C_i,\\ D_i,\\ A_i,\\ J_i \\mid X_i)$",
        "#EAF1F8", ec=C_MAIN, fs=10, weight="bold")
    box(ax, 7.2, 4.7, 5.5, 1.1, "跨命题文化结构\nCross-proposition Structure\n$S_g = Structure(U_1, U_2, \\dots, U_n)$",
        "#F0E8F0", ec="#7F6BAF", fs=10, weight="bold")
    arrow(ax, 6.8, 7.1, 7.2, 7.15, color=C_MAIN, lw=2)
    arrow(ax, 9.95, 6.6, 9.95, 5.8, color="#7F6BAF", lw=2)

    # 右下：说明
    box(ax, 7.2, 2.4, 5.5, 1.9,
        "方法论原则 Methodological principles\n\n"
        "• 单题只能作指标，不能单独证明图式\n  Single item = indicator, not a schema\n"
        "• 结构判断须基于多题及其关系\n  Structure needs multiple propositions\n"
        "• 国家是参照总体，非固定文化本体\n  Nation = reference population, not culture itself",
        "#FFF8EE", ec=C_ORANGE, fs=8.8)

    # 底部实例
    box(ax, 1.1, 0.3, 11.6, 1.5,
        "实例 Example：WVS Q29「男性总体上比女性更适合担任政治领导者」\n"
        "→ C: 平等—等级 | D: 政治法律 D-PL | A/O: 性别 O-GC + 政治角色 O-AR | J: 评价排序 J-ER | X: 元数据未观测 C-UNOBS",
        "#F7F7F7", ec=C_GREY, fs=9, weight="bold")
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig07_framework_architecture.png"), bbox_inches="tight")
    plt.close()

# ================= 图8：编码流程图（四层工作链） =================
def fig08():
    fig, ax = plt.subplots(figsize=(13,9.5))
    ax.set_xlim(0,13); ax.set_ylim(0,10); ax.axis("off")
    ax.set_title("图8  编码体系设计工作链：从理论框架到文化图式的循序渐进\n"
                "Fig.8  Encoding Pipeline: from Framework to Cultural Schema",
                fontsize=13, fontweight="bold", pad=10)

    stages = [
        (0.5, "① 方法论层\nMethodology", "说明文档.docx\n核心框架.xlsx", "6层框架 C/D/A/J/S/X\n定义、原则、编码实例", C_MAIN),
        (3.2, "② 编码规范层\nCoding Spec", "语义编号表.xlsx", "V/D/O/J/C/S 代码字典\n65条标准化代码+边界定义", C_GREEN),
        (5.9, "③ 数据执行层\nData Execution", "wvs编码结果表.xlsx", "WVS全290题逐题编码\n原子命题+多维打码+裁决审计", C_ORANGE),
        (8.6, "④ 理论提升层\nTheory Lift", "中层理论注册表.xlsx", "34条中层理论\nJODCV模式→理论机制→证据/反证", C_ACC),
    ]
    for x, title, fname, desc, col in stages:
        # 顶部标题卡
        box(ax, x, 7.6, 2.4, 1.1, title, col, ec=col, tc="white", fs=10, weight="bold")
        # 文件卡
        box(ax, x, 5.9, 2.4, 1.2, fname, "#FFFFFF", ec=col, fs=9.5, weight="bold")
        # 描述卡
        box(ax, x, 3.5, 2.4, 2.1, desc, "#FAFAFA", ec=col, fs=9)
        # 箭头 标题→文件→描述
        arrow(ax, x+1.2, 7.6, x+1.2, 7.1, color=col, lw=1.5)
        arrow(ax, x+1.2, 5.9, x+1.2, 5.6, color=col, lw=1.5)

    # 阶段间横向箭头
    for i in range(3):
        x1 = stages[i][0]+2.4; x2 = stages[i+1][0]
        arrow(ax, x1, 8.15, x2, 8.15, color=C_GREY, lw=2.2)
        # 标注转化动作
        actions = ["转写为可执行代码\nOperationalize","逐题打码\nItem coding","按模式上升\nPattern lift"]
        ax.text((x1+x2)/2, 8.45, actions[i], ha="center", fontsize=8, color=C_GREY, style="italic")

    # 底部：第五步 文化图式（目标）
    box(ax, 8.6, 0.8, 2.4, 1.8, "⑤ 结构判断层\nStructure Judgment\n文化图式 S-SC\nCultural Schema", "#7F6BAF", ec="#7F6BAF", tc="white", fs=9.5, weight="bold")
    arrow(ax, 9.8, 3.5, 9.8, 2.6, color="#7F6BAF", lw=2.2)
    ax.text(10.1, 3.05, "多题联合检验\nJoint test", fontsize=8, color="#7F6BAF", style="italic")

    # 左下：当前进展
    box(ax, 0.5, 0.8, 7.6, 1.8,
        "当前进展 Current Progress\n"
        "● 框架已建 ● 代码字典已定 ● 290题全量编码 ● 34理论已注册\n"
        "→ 下一阶段：人工复核 + 题项群组织为 S1/S2/S3/S4 结构单元 → 提炼文化图式",
        "#FFF8EE", ec=C_ORANGE, fs=9.5)
    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig08_encoding_pipeline.png"), bbox_inches="tight")
    plt.close()

# ================= 图9：理论→图式上升图 =================
def fig09():
    fig, ax = plt.subplots(figsize=(14,10))
    ax.set_xlim(0,14); ax.set_ylim(0,10.5); ax.axis("off")
    ax.set_title("图9  从原子命题到文化图式：编码→中层理论→图式的上升机制\n"
                "Fig.9  From Atomic Propositions to Cultural Schema (with real examples)",
                fontsize=13, fontweight="bold", pad=10)

    # ---- 第1层：290原子命题 ----
    box(ax, 0.3, 8.6, 13.4, 1.1,
        "第1层  290个原子命题  Layer 1  Atomic Propositions  (wvs编码结果表)\n"
        "每题 = 原子命题 + (J, O, D, C, V) 多维编码",
        C_MAIN, ec=C_MAIN, tc="white", fs=10, weight="bold")

    # ---- 第2层：具体题项（真实例子）----
    # 例1：Q29, Q28（性别）
    box(ax, 0.3, 6.9, 6.3, 1.3,
        "Q29 男性更适合当政治领导者\n  J-ER | O-GC+O-AR | D-PL\n"
        "Q28 母亲工作使儿童受损\n  J-AC | O-GC+O-FR | D-FI+D-WO",
        "#EAF1F8", ec=C_MAIN, fs=8.8)
    # 例2：Q57-Q61（信任）
    box(ax, 7.0, 6.9, 6.7, 1.3,
        "Q57-Q61 信任家人/邻居/熟人/陌生人\n  均为 J-RA | O-ST | D-CR\n"
        "→ 关系距离递增的信任序列",
        "#EAF1F8", ec=C_MAIN, fs=8.8)

    # ---- 第3层：JODCV模式分组 ----
    box(ax, 0.3, 5.2, 13.4, 1.1,
        "第2层  按理论注册表 JODCV_entry_S_form 模式匹配分组  Layer 2  Pattern-matching by JODCV",
        C_GREEN, ec=C_GREEN, tc="white", fs=10, weight="bold")

    # 模式标注
    box(ax, 0.3, 3.7, 6.3, 1.1,
        "匹配模式 Pattern：\nJ-ER/AC + O-GC + 跨D → 性别化判断群",
        "#EFF7F0", ec=C_GREEN, fs=9)
    box(ax, 7.0, 3.7, 6.7, 1.1,
        "匹配模式 Pattern：\nJ-RA + O-ST + D-CR → 关系距离信任群",
        "#EFF7F0", ec=C_GREEN, fs=9)

    arrow(ax, 3.45, 6.9, 3.45, 4.8, color=C_GREEN, lw=1.5)
    arrow(ax, 10.35, 6.9, 10.35, 4.8, color=C_GREEN, lw=1.5)

    # ---- 第4层：中层理论 ----
    box(ax, 0.3, 2.2, 6.3, 1.3,
        "MR18  「做性别」与性别化制度\n"
        "MR12  家庭奉献—工作奉献竞争图式",
        "#FDEEEC", ec=C_ACC, fs=9, weight="bold")
    box(ax, 7.0, 2.2, 6.7, 1.3,
        "MR02  象征边界及其社会后果\n"
        "MR15  互惠规范",
        "#FDEEEC", ec=C_ACC, fs=9, weight="bold")
    arrow(ax, 3.45, 3.7, 3.45, 3.5, color=C_ACC, lw=1.5)
    arrow(ax, 10.35, 3.7, 10.35, 3.5, color=C_ACC, lw=1.5)

    # ---- 第5层：文化图式 S-SC ----
    box(ax, 3.0, 0.4, 8.0, 1.3,
        "第3层  文化图式 S-SC  Cultural Schema  (可证伪结构)\n"
        "需满足：联合分布 + 概念关联网络 + 跨情境迁移（MR01 标准）",
        "#7F6BAF", ec="#7F6BAF", tc="white", fs=9.8, weight="bold")
    arrow(ax, 3.45, 2.2, 6.2, 1.7, color="#7F6BAF", lw=2)
    arrow(ax, 10.35, 2.2, 7.8, 1.7, color="#7F6BAF", lw=2)

    # 右侧关键说明
    box(ax, 11.0, 2.2, 2.7, 1.3,
        "单题≠图式\nSingle item ≠ schema\n须多题联合\nJoint evidence",
        "#FFF8EE", ec=C_ORANGE, fs=8.8)

    plt.tight_layout()
    plt.savefig(os.path.join(OUT,"fig09_theory_to_schema.png"), bbox_inches="tight")
    plt.close()

# ================= HTML 索引页 =================
def make_html():
    charts = [
        ("fig01_judgment_J.png","图1 290题×判断类型J 雷达图","A.统计分布图"),
        ("fig02_domain_D.png","图2 290题×场域D 雷达图","A.统计分布图"),
        ("fig03_object_O.png","图3 290题×对象O 雷达图","A.统计分布图"),
        ("fig04_theory_section.png","图4 34理论×板块 雷达图","A.统计分布图"),
        ("fig05_review_status.png","图5 review_status裁决状态分布","B.补充统计图"),
        ("fig06_JxD_heatmap.png","图6 J×D交叉热力图","B.补充统计图"),
        ("fig07_framework_architecture.png","图7 6层框架架构图","C.设计叙事图"),
        ("fig08_encoding_pipeline.png","图8 编码体系工作链","C.设计叙事图"),
        ("fig09_theory_to_schema.png","图9 理论→图式上升图","C.设计叙事图"),
    ]
    rows = "".join(
        f'<div class="card"><h3>{t}</h3><img src="{f}" loading="lazy" onclick="this.requestFullscreen&&this.requestFullscreen()"></div>'
        for f,t,_ in charts)
    html = f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<title>多元文化建模项目可视化图表</title>
<style>
body{{font-family:'PingFang SC','Helvetica',sans-serif;max-width:1200px;margin:0 auto;padding:24px;background:#fafafa;color:#222;}}
h1{{color:#2E5C8A;border-bottom:3px solid #2E5C8A;padding-bottom:10px;}}
.card{{background:#fff;border:1px solid #e0e0e0;border-radius:8px;padding:16px;margin:14px 0;box-shadow:0 1px 3px rgba(0,0,0,.06);}}
.card h3{{margin:0 0 12px;color:#2E5C8A;font-size:16px;}}
img{{width:100%;border:1px solid #eee;border-radius:4px;cursor:zoom-in;}}
.intro{{background:#EAF1F8;padding:14px 18px;border-radius:8px;font-size:14px;line-height:1.7;}}
</style></head><body>
<h1>多元文化建模项目 — 可视化图表索引</h1>
<div class="intro">
<b>项目</b>：跨文化知识建模 · WVS 290题编码体系<br>
<b>内容</b>：A 统计分布图(雷达图) → B 补充统计图 → C 设计叙事流程图<br>
<b>数据</b>：全部来自项目Excel真实数据，中英双语标签<br>
<b>用途</b>：学术汇报/答辩
</div>
{rows}
</body></html>"""
    with open(os.path.join(OUT,"index.html"),"w",encoding="utf-8") as f:
        f.write(html)

# ---------------- 主流程 ----------------
if __name__ == "__main__":
    print(f"加载：WVS {N_Q} 题，理论 {N_T} 条，语义 {len(SEM)} 行")
    fig01(); print("● fig01 J")
    fig02(); print("● fig02 D")
    fig03(); print("● fig03 O")
    fig04(); print("● fig04 theory section")
    fig05(); print("● fig05 review_status")
    fig06(); print("● fig06 J×D heatmap")
    fig07(); print("● fig07 framework")
    fig08(); print("● fig08 pipeline")
    fig09(); print("● fig09 theory→schema")
    make_html(); print("● index.html")
    print(f"\n全部完成，输出目录：{OUT}")
