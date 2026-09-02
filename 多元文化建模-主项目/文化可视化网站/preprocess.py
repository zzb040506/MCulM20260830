# -*- coding: utf-8 -*-
"""数据预处理管道：
1. 读取 WVS-7 完整 CSV + 290题编码表
2. 290题变量映射到 WVS 列(QN 或 QNP)
3. 每题 min-max 归一化(全局，跨全部受访者)
4. 按编码维度(D场域/J判断)聚合为受访者级维度分
5. 输出小表：受访者维度分 / 国家维度汇总 / 题项国家均值 / 国家元数据
"""
import os, sys, json
import numpy as np
import pandas as pd
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # 项目根 多元文化建模20260830
PROJ = os.path.join(BASE, "多元文化建模-主项目")
WVS_CSV = os.path.join(BASE, "WVS-7数据集", "WVS_Cross-National_Wave_7_inverted_csv_v6_0.csv")
OUT = os.path.join(PROJ, "文化可视化网站", "data")
os.makedirs(OUT, exist_ok=True)

sys.path.insert(0, os.path.join(PROJ, "文化可视化网站"))
from country_meta import COUNTRIES, zone_of, name_zh, name_en

# 维度中文映射
D_ZH = {"D-FI":"家庭","D-CR":"社区","D-RS":"宗教","D-PL":"政治法律","D-WO":"市场工作",
        "D-KT":"教育知识","D-MC":"媒介文化","D-HC":"健康身体"}
J_ZH = {"J-CM":"分类","J-AC":"属性因果","J-ER":"评价排序","J-NO":"规范应然",
        "J-RA":"关系分配","J-RI":"表征认同"}
# O 对象：按语义编号表顺序，仅保留实际在 O_primary 出现过的 14 个代码
O_ZH = {"O-GC":"性别身份 Gender","O-SF":"性/家庭形式 Sex&Family","O-AG":"年龄世代 Age",
        "O-SP":"社会经济位置 Socioecon","O-RE":"种族族群 Race","O-NM":"国族迁移 Nation&Migr",
        "O-RS":"宗教世俗身份 Religious","O-LG":"语言群体 Language","O-BH":"身体健康 Body&Health",
        "O-FR":"家庭角色 Family Roles","O-ST":"社会关系 Social Ties","O-AR":"权威角色 Authority",
        "O-GP":"群体公众 Groups","O-OI":"组织机构 Org&Inst","O-PL":"实践生活方式 Practices",
        "O-RP":"规则政策 Rules","O-KT":"知识技术 Knowledge&Tech","O-MSW":"媒介符号 Media",
        "O-OSN":"物空自然 Objects&Space","O-OPEN":"开放对象 Open Object"}
O_ORDER = ["O-GC","O-SF","O-SP","O-NM","O-RS","O-BH","O-FR","O-ST",
           "O-AR","O-GP","O-OI","O-PL","O-RP","O-KT"]
D_ORDER = ["D-FI","D-CR","D-RS","D-PL","D-WO","D-KT","D-MC","D-HC"]
J_ORDER = ["J-CM","J-AC","J-ER","J-NO","J-RA","J-RI"]

def load_encoding():
    wb = openpyxl.load_workbook(os.path.join(PROJ, "国家文化知识建模_wvs编码结果表.xlsx"), data_only=True)
    ws = wb.active
    h = [c.value for c in ws[1]]
    rows = [dict(zip(h, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    enc = pd.DataFrame(rows)
    # 变量映射函数
    return enc

def main():
    print("① 读取编码表与WVS列名...")
    enc = load_encoding()
    # 读 WVS 表头
    import csv
    with open(WVS_CSV) as f:
        wvs_cols = set(next(csv.reader(f)))

    def mapcol(v):
        if v in wvs_cols: return v
        if (v+"P") in wvs_cols: return v+"P"
        return None
    enc["wvs_col"] = enc["variable"].map(mapcol)
    assert enc["wvs_col"].notna().all(), enc[enc["wvs_col"].isna()]
    print(f"   290题全部映射成功，{len(enc)}题")

    # 仅取内容编码题（J/D/O 非 —）
    content = enc[~enc["J_code"].isin(["—", None]) & ~enc["D_primary"].isin(["—", None, "OPEN-PROVISIONAL"])].copy()
    print(f"   内容编码题 {len(content)}/{len(enc)}（去—与provisional）")

    print("② 读取WVS数据(pandas)...")
    need = ["B_COUNTRY","B_COUNTRY_ALPHA","Q260","Q262","Q275R","Q288R","SACSECVAL","RESEMAVAL"]
    need += list(enc["wvs_col"].unique())
    df = pd.read_csv(WVS_CSV, usecols=lambda c: c in need, low_memory=False)
    print(f"   {df.shape[0]} 行 × {df.shape[1]} 列")

    print("③ 数值化 + 缺失值(负值)处理 + min-max归一化...")
    qcols = list(enc["wvs_col"].unique())
    for c in qcols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
        df.loc[df[c] < 0, c] = np.nan
    norm_cols = []
    for c in qcols:
        s = df[c]
        mn, mx = s.min(), s.max()
        nc = c + "__n"
        if pd.isna(mn) or mx == mn:
            df[nc] = 0.5
        else:
            df[nc] = (s - mn) / (mx - mn)
        norm_cols.append(nc)

    print("④ 计算受访者级维度分(按D/J聚合各题归一化分)...")
    # D 维度分
    dim_cols = []
    for d in D_ORDER:
        sub = content[content["D_primary"] == d]
        cols = [c + "__n" for c in sub["wvs_col"]]
        cols = [c for c in cols if c in df.columns]
        if cols:
            df["dim_" + d] = df[cols].mean(axis=1, skipna=True)
            dim_cols.append("dim_" + d)
    for j in J_ORDER:
        sub = content[content["J_code"] == j]
        cols = [c + "__n" for c in sub["wvs_col"]]
        cols = [c for c in cols if c in df.columns]
        if cols:
            df["dim_" + j] = df[cols].mean(axis=1, skipna=True)
            dim_cols.append("dim_" + j)
    # O 对象维度聚合（按 O_primary）
    for o in O_ORDER:
        sub = content[content["O_primary"] == o]
        cols = [c + "__n" for c in sub["wvs_col"]]
        cols = [c for c in cols if c in df.columns]
        if cols:
            df["dim_" + o] = df[cols].mean(axis=1, skipna=True)
            dim_cols.append("dim_" + o)
    print(f"   生成维度列: {dim_cols}")

    print("⑤ 处理人口学与Welzel指数...")
    df["sex"] = df["Q260"].map({1.0: "男 Male", 2.0: "女 Female"})
    df["age"] = pd.to_numeric(df["Q262"], errors="coerce")
    df["age_group"] = pd.cut(df["age"], [0, 30, 45, 60, 200],
                             labels=["<30", "30-44", "45-59", "60+"])
    df["edu"] = df["Q275R"].map({1.0: "低 Low", 2.0: "中 Mid", 3.0: "高 High"})
    df["income"] = df["Q288R"].map({1.0: "低 Low", 2.0: "中 Mid", 3.0: "高 High"})
    df["country"] = df["B_COUNTRY_ALPHA"]
    df["zone"] = df["country"].map(zone_of)
    # Welzel: 检查量纲
    for w in ["SACSECVAL", "RESEMAVAL"]:
        df[w] = pd.to_numeric(df[w], errors="coerce")
        df.loc[df[w] < 0, w] = np.nan
        if df[w].max() > 1.5:  # 0-100 量纲
            df[w] = df[w] / 100.0
    df["welzel_secular"] = df["SACSECVAL"]    # 传统→世俗理性(高=世俗)
    df["welzel_emanc"] = df["RESEMAVAL"]      # 生存→自我表达(高=解放)

    print("⑥ 保存受访者维度表(供人口学切片实时聚合)...")
    out_cols = ["country", "zone", "sex", "age_group", "edu", "income",
                "welzel_secular", "welzel_emanc"] + dim_cols
    resp = df[out_cols].copy()
    resp.to_parquet(os.path.join(OUT, "respondent_dims.parquet"))
    print(f"   respondent_dims.parquet: {resp.shape}")

    print("⑦ 保存国家维度汇总(全样本)...")
    agg = resp.groupby("country")[["welzel_secular", "welzel_emanc"] + dim_cols].mean()
    n = resp.groupby("country").size().rename("n_resp")
    agg = agg.join(n).reset_index()
    agg["name_zh"] = agg["country"].map(name_zh)
    agg["name_en"] = agg["country"].map(name_en)
    agg["zone"] = agg["country"].map(zone_of)
    agg.to_parquet(os.path.join(OUT, "country_summary.parquet"), index=False)
    print(f"   country_summary.parquet: {agg.shape}")

    print("⑧ 保存题项国家均值(题项级页)...")
    item = df.groupby("country")[norm_cols].mean()
    # 列名还原为 Q编号
    colmap = {c + "__n": c for c in qcols}
    # 反查 Q编号
    enc_map = dict(zip(enc["wvs_col"], enc["variable"]))
    item.columns = [enc_map.get(c.replace("__n", ""), c) for c in item.columns]
    item = item.reset_index()
    # 附题目文字
    prop_map = dict(zip(enc["variable"], enc["atomic_proposition"]))
    jmap = dict(zip(enc["variable"], enc["J_code"]))
    dmap = dict(zip(enc["variable"], enc["D_primary"]))
    item.to_parquet(os.path.join(OUT, "item_means.parquet"), index=False)
    # 题项元数据单独存
    item_meta = enc[["variable", "wvs_col", "atomic_proposition", "J_code", "O_primary",
                     "D_primary", "C_codes", "V_primary_direction", "review_status"]].copy()
    item_meta.to_parquet(os.path.join(OUT, "item_meta.parquet"), index=False)
    print(f"   item_means.parquet: {item.shape}, item_meta.parquet: {item_meta.shape}")

    print("\n✅ 全部预处理完成，输出目录:", OUT)
    print("   文件:", os.listdir(OUT))

if __name__ == "__main__":
    main()
