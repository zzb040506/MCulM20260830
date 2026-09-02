# -*- coding: utf-8 -*-
"""共享工具：数据加载、筛选、维度配置、配色"""
import os, sys
import pandas as pd
import numpy as np
import streamlit as st

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from country_meta import COUNTRIES, ZONES, ZONE_EN, ZONE_COLOR, name_zh, name_en, zone_of

DATA = os.path.join(HERE, "data")

# 维度配置
D_ORDER = ["D-FI","D-CR","D-RS","D-PL","D-WO","D-KT","D-MC","D-HC"]
J_ORDER = ["J-CM","J-AC","J-ER","J-NO","J-RA","J-RI"]
O_ORDER = ["O-GC","O-SF","O-SP","O-NM","O-RS","O-BH","O-FR","O-ST",
           "O-AR","O-GP","O-OI","O-PL","O-RP","O-KT"]
D_ZH = {"D-FI":"家庭 Family","D-CR":"社区 Community","D-RS":"宗教 Religion",
        "D-PL":"政治法律 Polity","D-WO":"市场工作 Work","D-KT":"教育知识 Knowledge",
        "D-MC":"媒介文化 Media","D-HC":"健康身体 Health"}
J_ZH = {"J-CM":"分类 Classify","J-AC":"属性因果 Attribute","J-ER":"评价排序 Evaluate",
        "J-NO":"规范应然 Norm","J-RA":"关系分配 Relate","J-RI":"表征认同 Represent"}
O_ZH = {"O-GC":"性别身份 Gender","O-SF":"性/家庭形式 Sex&Family","O-AG":"年龄世代 Age",
        "O-SP":"社会经济位置 Socioecon","O-RE":"种族族群 Race","O-NM":"国族迁移 Nation&Migr",
        "O-RS":"宗教世俗身份 Religious","O-LG":"语言群体 Language","O-BH":"身体健康 Body&Health",
        "O-FR":"家庭角色 Family Roles","O-ST":"社会关系 Social Ties","O-AR":"权威角色 Authority",
        "O-GP":"群体公众 Groups","O-OI":"组织机构 Org&Inst","O-PL":"实践生活方式 Practices",
        "O-RP":"规则政策 Rules","O-KT":"知识技术 Knowledge&Tech","O-MSW":"媒介符号 Media",
        "O-OSN":"物空自然 Objects&Space","O-OPEN":"开放对象 Open Object"}
DIMS = {"D": ("制度与生活场域 Domain D", D_ORDER, D_ZH),
        "J": ("判断与关系规则 Judgment J", J_ORDER, J_ZH),
        "O": ("文化判断对象 Object O", O_ORDER, O_ZH)}

@st.cache_data
def load_country_summary():
    return pd.read_parquet(os.path.join(DATA, "country_summary.parquet"))

@st.cache_data
def load_respondent_dims():
    return pd.read_parquet(os.path.join(DATA, "respondent_dims.parquet"))

@st.cache_data
def load_item_means():
    return pd.read_parquet(os.path.join(DATA, "item_means.parquet"))

@st.cache_data
def load_item_meta():
    return pd.read_parquet(os.path.join(DATA, "item_meta.parquet"))

@st.cache_data
def load_theory():
    # 优先用打包的 parquet（云端部署/无 Excel 时）；
    # 本地开发回退到原 Excel（保留 openpyxl 行为）。
    pq = os.path.join(DATA, "theory.parquet")
    if os.path.exists(pq):
        return pd.read_parquet(pq)
    import openpyxl
    PARENT = os.path.dirname(HERE)  # 多元文化建模-主项目
    xlsx = os.path.join(PARENT, "国家文化知识建模_中层理论注册表.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    h = [c.value for c in ws[1]]
    rows = [dict(zip(h, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return pd.DataFrame(rows)

def dim_cols(kind):
    return [f"dim_{c}" for c in (D_ORDER if kind=="D" else J_ORDER if kind=="J" else O_ORDER)]

def country_summary_filtered(demo):
    """根据人口学筛选返回国家×维度汇总。demo: dict(sex/age/edu/income)。"""
    has_filter = any(demo.get(k) for k in ["sex","age","edu","income"])
    base = load_country_summary().copy()
    name_map = dict(zip(base["country"], base["name_zh"]))
    nameen_map = dict(zip(base["country"], base["name_en"]))
    if not has_filter:
        return base
    rd = load_respondent_dims()
    m = pd.Series(True, index=rd.index)
    if demo.get("sex"):   m &= (rd["sex"]==demo["sex"])
    if demo.get("age"):   m &= (rd["age_group"].astype(str)==demo["age"])
    if demo.get("edu"):   m &= (rd["edu"]==demo["edu"])
    if demo.get("income"):m &= (rd["income"]==demo["income"])
    sub = rd[m]
    cols = ["welzel_secular","welzel_emanc"] + dim_cols("D") + dim_cols("J") + dim_cols("O")
    agg = sub.groupby("country")[cols].mean().reset_index()
    agg["n_resp"] = sub.groupby("country").size().reindex(agg["country"]).values
    agg["name_zh"] = agg["country"].map(name_map)
    agg["name_en"] = agg["country"].map(nameen_map)
    agg["zone"] = agg["country"].map(zone_of)
    return agg

def render_sidebar():
    """渲染侧边栏筛选控件，返回 dict: zones, dim_kind, demo."""
    st.sidebar.markdown("## 🔧 筛选 Filters")
    zones = st.sidebar.multiselect("文化圈 Cultural zone", ZONES, default=ZONES,
                                   format_func=lambda z: f"{z} / {ZONE_EN[z]}")
    dim_kind = st.sidebar.radio("维度集 Dimension set", ["D","J","O"], format_func=lambda k: DIMS[k][0])
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 👤 人口学切片 Demographics")
    demo = {}
    demo["sex"] = st.sidebar.selectbox("性别 Sex", ["", "男 Male", "女 Female"]) or None
    demo["age"] = st.sidebar.selectbox("年龄 Age", ["", "<30","30-44","45-59","60+"]) or None
    demo["edu"] = st.sidebar.selectbox("教育 Education", ["", "低 Low","中 Mid","高 High"]) or None
    demo["income"] = st.sidebar.selectbox("收入 Income", ["", "低 Low","中 Mid","高 High"]) or None
    return {"zones": zones, "dim_kind": dim_kind, "demo": demo}

def filter_countries(df, zones):
    if zones:
        return df[df["zone"].isin(zones)].copy()
    return df.copy()

def fmt_dim(code, kind):
    return DIMS[kind][2].get(code, code)

# 题项文字（中文简化）
@st.cache_data
def question_text(qid):
    im = load_item_meta()
    row = im[im["variable"]==qid]
    if len(row)==0: return ""
    return str(row.iloc[0]["atomic_proposition"])
