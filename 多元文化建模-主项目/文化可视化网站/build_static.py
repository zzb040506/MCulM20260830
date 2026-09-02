# -*- coding: utf-8 -*-
"""导出静态站数据 docs/data.js（含人口学切片）"""
import os, sys, json
import numpy as np
import pandas as pd
import openpyxl

BASE = "/Users/f.fantasiachopin/Documents/UCAS博士文件夹/Project/多元文化建模20260830"
PROJ = os.path.join(BASE, "多元文化建模-主项目")
APP = os.path.join(PROJ, "文化可视化网站")
DATA = os.path.join(APP, "data")
DOCS = os.path.join(BASE, "docs")
os.makedirs(DOCS, exist_ok=True)
sys.path.insert(0, APP)
from country_meta import COUNTRIES, ZONES, ZONE_EN, ZONE_COLOR, name_zh, name_en, zone_of

D_ORDER = ["D-FI","D-CR","D-RS","D-PL","D-WO","D-KT","D-MC","D-HC"]
J_ORDER = ["J-CM","J-AC","J-ER","J-NO","J-RA","J-RI"]
D_ZH = {"D-FI":"家庭","D-CR":"社区","D-RS":"宗教","D-PL":"政治法律","D-WO":"市场工作",
        "D-KT":"教育知识","D-MC":"媒介文化","D-HC":"健康身体"}
J_ZH = {"J-CM":"分类","J-AC":"属性因果","J-ER":"评价排序","J-NO":"规范应然",
        "J-RA":"关系分配","J-RI":"表征认同"}
DIM_COLS = [f"dim_{c}" for c in D_ORDER+J_ORDER]
WELZEL = ["welzel_secular","welzel_emanc"]

def country_records(df):
    """df: 受访者级(已筛选) → 国家汇总记录列表"""
    g = df.groupby("country")
    agg = g[WELZEL+DIM_COLS].mean().reset_index()
    agg["n_resp"] = g.size().reindex(agg["country"]).values
    recs = []
    for _, r in agg.iterrows():
        iso = r["country"]
        rec = {
            "iso3": iso, "name_zh": name_zh(iso), "name_en": name_en(iso),
            "zone": zone_of(iso), "n_resp": int(r["n_resp"]),
            "welzel_secular": round(float(r["welzel_secular"]),4) if pd.notna(r["welzel_secular"]) else None,
            "welzel_emanc": round(float(r["welzel_emanc"]),4) if pd.notna(r["welzel_emanc"]) else None,
            "color": ZONE_COLOR.get(zone_of(iso),"#888"),
        }
        for c in DIM_COLS:
            v = r[c]
            rec[c] = round(float(v),4) if pd.notna(v) else None
        recs.append(rec)
    return recs

def main():
    rd = pd.read_parquet(os.path.join(DATA, "respondent_dims.parquet"))
    # age_group 转字符串便于比较
    rd["age_group"] = rd["age_group"].astype(str)
    rd["sex"] = rd["sex"].astype(str)
    rd["edu"] = rd["edu"].astype(str)
    rd["income"] = rd["income"].astype(str)

    slices = {"all": country_records(rd)}
    for v in ["男 Male","女 Female"]:
        slices[f"sex:{v}"] = country_records(rd[rd["sex"]==v])
    for a in ["<30","30-44","45-59","60+"]:
        slices[f"age:{a}"] = country_records(rd[rd["age_group"]==a])
    for e in ["低 Low","中 Mid","高 High"]:
        slices[f"edu:{e}"] = country_records(rd[rd["edu"]==e])
    for inc in ["低 Low","中 Mid","高 High"]:
        slices[f"income:{inc}"] = country_records(rd[rd["income"]==inc])

    # item_means: country × question (normalized 0-1 means)
    im = pd.read_parquet(os.path.join(DATA, "item_means.parquet"))
    im_meta = pd.read_parquet(os.path.join(DATA, "item_meta.parquet"))
    # 题项→元数据
    meta_map = {row["variable"]: row for _,row in im_meta.iterrows()}
    item_recs = []
    qcols = [c for c in im.columns if c != "country"]
    for _, row in im.iterrows():
        v = row["country"]
        rec = {"variable": v, "values": {}}
        for q in qcols:
            m = im_meta[im_meta["variable"]==q]
            if len(m):
                # 该国该题均值，按 question 列名
                pass
        item_recs.append({"country": v, "name_zh": name_zh(v)})
    # 重组为 题项→{country: mean}
    item_by_q = {}
    for q in qcols:
        rec = {"variable": q,
               "text": str(meta_map.get(q, pd.Series()).get("atomic_proposition","")) if q in meta_map else "",
               "J": str(meta_map.get(q, pd.Series()).get("J_code","")) if q in meta_map else "",
               "D": str(meta_map.get(q, pd.Series()).get("D_primary","")) if q in meta_map else ""}
        vals = {}
        for _, row in im.iterrows():
            c = row["country"]
            val = row[q]
            vals[c] = round(float(val),4) if pd.notna(val) else None
        rec["values"] = vals
        item_by_q[q] = rec
    # 仅保留内容编码题(去—)
    item_list = [item_by_q[q] for q in qcols
                 if item_by_q[q]["J"] not in ("—","None","nan","") and item_by_q[q]["D"] not in ("—","None","nan","")]

    # theory
    wb = openpyxl.load_workbook(os.path.join(PROJ, "国家文化知识建模_中层理论注册表.xlsx"), data_only=True)
    ws = wb.active; h=[c.value for c in ws[1]]
    th = [dict(zip(h,r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    theory = []
    for r in th:
        theory.append({
            "id": r.get("theory_id"), "name": r.get("theory_name"), "section": r.get("section"),
            "source": r.get("source_level"),
            "mechanism": str(r.get("core_mechanism_expected_relation",""))[:300],
            "jodcv": str(r.get("JODCV_entry_S_form","")),
            "evidence": str(r.get("observable_evidence_falsification",""))[:300],
            "literature": str(r.get("literature_basis",""))[:200],
        })

    # 维度构成统计
    from collections import Counter
    jcomp = [{"code":k,"n":v,"label":J_ZH.get(k,k)} for k,v in Counter(im_meta["J_code"]).items() if k not in ("—",None)]
    dcomp = [{"code":k,"n":v,"label":D_ZH.get(k,k)} for k,v in Counter(im_meta["D_primary"]).items() if k not in ("—",None,"OPEN-PROVISIONAL")]

    out = {
        "slices": slices,
        "items": item_list,
        "theories": theory,
        "zones": [{"zone":z,"en":ZONE_EN[z],"color":ZONE_COLOR[z]} for z in ZONES],
        "dimConfig": {
            "D": {"order":D_ORDER, "labels":[{c:D_ZH[c]} for c in D_ORDER]},
            "J": {"order":J_ORDER, "labels":[{c:J_ZH[c]} for c in J_ORDER]},
        },
        "composition": {"J":jcomp, "D":dcomp},
    }
    with open(os.path.join(DOCS,"data.js"),"w",encoding="utf-8") as f:
        f.write("window.APP_DATA=")
        f.write(json.dumps(out, ensure_ascii=False))
        f.write(";")
    print(f"data.js 写入完成: {os.path.getsize(os.path.join(DOCS,'data.js'))//1024} KB")
    print(f"切片数: {len(slices)}, 题项数: {len(item_list)}, 理论数: {len(theory)}")

if __name__ == "__main__":
    main()
